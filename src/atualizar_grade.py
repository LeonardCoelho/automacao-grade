import pandas as pd
from openpyxl import load_workbook, Workbook
from datetime import datetime, timedelta
from copy import copy
import os
import traceback
import subprocess
import time
import win32com.client
import pythoncom

def planilha_esta_aberta(nome_arquivo_excel):
    """Verifica se o Excel tem o arquivo aberto (tratando nomes temporários)."""
    pythoncom.CoInitialize()
    try:
        excel = win32com.client.GetActiveObject("Excel.Application")
        for wb in excel.Workbooks:
            nome_aberto = wb.Name.lower().replace("~$", "")
            if nome_arquivo_excel.lower().endswith(nome_aberto):
                return True
    except Exception:
        pass
    return False

try:
    print("🚀 Iniciando script...")

    caminho_original = r"C:\Users\leosouza\Arcor\GC_BR_Planejamento_Logístico - Documentos\3. BRA-Roteirizacao\Ordens de Coleta 2025\05 CD CTG\ORDEM DE COLETA CTG.xlsx"
    nome_arquivo = os.path.basename(caminho_original)
    print(f"📂 Arquivo: {nome_arquivo}")

    # Espera até o arquivo ser liberado
    print("⏳ Aguardando o arquivo ser liberado...")
    mensagem_mostrada = False
    while True:
        try:
            with open(caminho_original, "rb") as f:
                f.read(1)
            break
        except PermissionError:
            if not mensagem_mostrada:
                print("🚫 O arquivo está em uso. Feche no Excel para continuar...")
                mensagem_mostrada = True
            time.sleep(3)

    print("✅ Arquivo liberado! Continuando...")

    # Define data da aba (pula domingo)
    hoje = datetime.today()
    amanha = hoje + timedelta(days=1)
    if amanha.weekday() == 6:  # Domingo
        amanha += timedelta(days=1)
    dia_planilha = str(amanha.day)
    print(f"🔎 Processando aba: '{dia_planilha}'")

    # Abre a planilha
    wb_data = load_workbook(caminho_original, data_only=True)
    ws_data = wb_data[dia_planilha]
    wb_formatos = load_workbook(caminho_original, data_only=False)
    ws_formatos = wb_formatos[dia_planilha]

    # Cria nova planilha
    wb_novo = Workbook()
    ws_novo = wb_novo.active
    ws_novo.title = dia_planilha

    # Copia cabeçalhos da linha 15
    for col_idx, cell in enumerate(ws_data[15], start=1):
        new_cell = ws_novo.cell(row=1, column=col_idx, value=cell.value)
        ref = ws_formatos.cell(row=15, column=col_idx)
        new_cell.font = copy(ref.font)
        new_cell.border = copy(ref.border)
        new_cell.fill = copy(ref.fill)
        new_cell.number_format = copy(ref.number_format)
        new_cell.protection = copy(ref.protection)
        new_cell.alignment = copy(ref.alignment)

    # Mapeia colunas
    headers = [str(cell.value).strip().upper() for cell in ws_data[15]]
    col_transp = headers.index("TRANSP.") + 1
    col_emailok = headers.index("EMAIL OK") + 1

    row_dest = 2
    for row_idx in range(16, ws_data.max_row + 1):
        transp = str(ws_data.cell(row=row_idx, column=col_transp).value).strip().upper()
        email_ok = str(ws_data.cell(row=row_idx, column=col_emailok).value).strip().upper()

        if transp == "MVC" and email_ok not in ["SEPARAR", "FATURAR"]:
            for col_idx in range(1, ws_data.max_column + 1):
                val = ws_data.cell(row=row_idx, column=col_idx).value
                new_cell = ws_novo.cell(row=row_dest, column=col_idx, value=val)
                ref = ws_formatos.cell(row=row_idx, column=col_idx)
                new_cell.font = copy(ref.font)
                new_cell.border = copy(ref.border)
                new_cell.fill = copy(ref.fill)
                new_cell.number_format = copy(ref.number_format)
                new_cell.protection = copy(ref.protection)
                new_cell.alignment = copy(ref.alignment)
            row_dest += 1

    # Ajuste de largura das colunas
    for col_letter, dim in ws_formatos.column_dimensions.items():
        ws_novo.column_dimensions[col_letter].width = dim.width

    # Ajuste da altura das linhas
    for row_idx in range(15, ws_formatos.max_row + 1):
        if row_idx in ws_formatos.row_dimensions:
            height = ws_formatos.row_dimensions[row_idx].height
            if height:
                ws_novo.row_dimensions[row_idx - 14].height = height

    # Salva o novo arquivo
    pasta_saida = r"C:\Users\leosouza\OneDrive - Arcor\Documentos\PowerAutomate"
    os.makedirs(pasta_saida, exist_ok=True)
    nome_saida = f"Grade_MVC_{dia_planilha}.xlsx"
    caminho_saida = os.path.join(pasta_saida, nome_saida)
    caminho_temp = os.path.join(pasta_saida, f"__temp_{nome_saida}")

    wb_novo.save(caminho_temp)
    os.replace(caminho_temp, caminho_saida)

    print(f"✅ Arquivo final salvo com sucesso em: {caminho_saida}")

except Exception as e:
    erro_path = r"C:\Users\leosouza\erro_script.txt"
    with open(erro_path, "w") as f:
        f.write("Erro ao executar script:\n")
        f.write(traceback.format_exc())
    print(f"❌ ERRO AO EXECUTAR. Detalhes salvos em: {erro_path}")

finally:
    try:
        subprocess.Popen(f'start "" "excel" "{caminho_original}"', shell=True)
    except Exception as e:
        print("⚠️ Falha ao tentar reabrir o Excel:", e)

    input("🔚 Pressione ENTER para finalizar...")
